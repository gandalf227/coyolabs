import sys
import os
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, BASE_DIR)

import re
from typing import Dict, List, Optional, Tuple

import openpyxl


from app import create_app
from app.extensions import db
from app.models.lab import Lab
from app.models.material import Material




# Carpeta donde estarán los excels
INVENTORY_DIR = os.path.join("data", "inventarios")

# Palabras clave para detectar la fila de encabezados
HEADER_KEYWORDS = ["EQUIPO", "MATERIAL", "UBIC", "PZAS", "PIEZAS", "ESTADO", "OBS"]

# Normalización de nombres de columnas -> campo canónico
# (vamos a mapear por "contiene" para tolerar encabezados distintos)
COLUMN_RULES = {
    "name": ["EQUIPO", "MATERIAL", "EQUIPO/MATERIAL", "EQUIPO - MATERIAL"],
    "location": ["UBIC", "UBICACIÓN", "UBICACION", "ESTANTE", "GABINETE", "LUGAR"],
    "pieces_text": ["PZAS", "PIEZAS", "CANTIDAD", "NO. PIEZAS", "N° PIEZAS", "PIEZAS TOTALES"],
    "status": ["ESTADO", "CONDICION", "CONDICIÓN"],
    "brand": ["MARCA"],
    "model": ["MODELO"],
    "code": ["CODIGO", "CÓDIGO", "CLAVE"],
    "serial": ["SERIE", "NO. SERIE", "N° SERIE"],
    "notes": ["OBS", "OBSERVACIONES", "COMENTARIOS", "NOTAS"],
    "image_ref": ["FOTO", "IMAGEN", "EVIDENCIA"],
    "tutorial_url": ["TUTORIAL", "ENLACE", "URL"],
}


def normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def detect_header_row(rows: List[List[str]]) -> Optional[int]:
    """
    Encuentra la fila que parece encabezado, buscando varias palabras clave.
    """
    for i, row in enumerate(rows):
        joined = " ".join([normalize_header(c) for c in row if c is not None])
        hits = sum(1 for kw in HEADER_KEYWORDS if kw in joined)
        # Umbral conservador
        if hits >= 2:
            return i
    return None


def build_column_map(header_row: List[str]) -> Dict[str, int]:
    """
    Regresa un mapa: campo_canónico -> índice de columna
    basado en coincidencia parcial.
    """
    header_norm = [normalize_header(h) for h in header_row]
    mapping: Dict[str, int] = {}

    for field, candidates in COLUMN_RULES.items():
        for idx, h in enumerate(header_norm):
            for cand in candidates:
                cand_norm = normalize_header(cand)
                if cand_norm and cand_norm in h:
                    mapping[field] = idx
                    break
            if field in mapping:
                break

    return mapping


def get_cell(row: List, idx: int) -> Optional[str]:
    if idx < 0 or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    val = str(val).strip()
    return val if val != "" else None


def parse_pieces_qty(pieces_text: Optional[str]) -> Optional[int]:
    """
    Si pieces_text es "20" devuelve 20.
    Si es "20/20" devuelve 20 (primera parte).
    Si no se puede, devuelve None.
    """
    if not pieces_text:
        return None

    t = pieces_text.strip()
    # Ej: "20/20"
    if "/" in t:
        t = t.split("/")[0].strip()

    # Extraer primer número entero
    m = re.search(r"\d+", t)
    if not m:
        return None

    try:
        return int(m.group(0))
    except:
        return None


def ensure_lab(lab_name: str) -> Lab:
    lab_name = lab_name.strip()
    lab = Lab.query.filter_by(name=lab_name).first()
    if lab:
        return lab
    lab = Lab(name=lab_name)
    db.session.add(lab)
    db.session.commit()
    return lab


def derive_lab_name_from_filename(filename: str) -> str:
    """
    Intenta extraer el nombre del laboratorio del nombre del archivo.
    Ej: 'Inv - Lab - Redes y Soporte - 2024.xlsx' -> 'Redes y Soporte'
    """
    name = filename.replace(".xlsx", "").strip()
    parts = [p.strip() for p in name.split("-")]
    # Esperado: Inv / Lab / <LAB NAME> / 2024
    if len(parts) >= 3:
        return parts[2]
    return name


def import_workbook(path: str, lab: Lab) -> Tuple[int, int]:
    """
    Devuelve (insertados, saltados)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    inserted = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Leer primeras N filas como muestra para detectar encabezado
        preview_rows: List[List] = []
        max_preview = min(ws.max_row, 50)

        for r in range(1, max_preview + 1):
            preview_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

        header_idx = detect_header_row(preview_rows)
        if header_idx is None:
            # Hoja sin estructura esperada
            continue

        header_row = preview_rows[header_idx]
        col_map = build_column_map(header_row)

        # Para importar necesitamos al menos name (equipo/material)
        if "name" not in col_map:
            continue

        # Data empieza después del header
        start_row = header_idx + 2  # 1-indexed en openpyxl

        for r in range(start_row, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

            name = get_cell(row_vals, col_map["name"])
            if not name:
                skipped += 1
                continue

            location = get_cell(row_vals, col_map.get("location", -1)) if "location" in col_map else None
            status = get_cell(row_vals, col_map.get("status", -1)) if "status" in col_map else None
            pieces_text = get_cell(row_vals, col_map.get("pieces_text", -1)) if "pieces_text" in col_map else None

            brand = get_cell(row_vals, col_map.get("brand", -1)) if "brand" in col_map else None
            model = get_cell(row_vals, col_map.get("model", -1)) if "model" in col_map else None
            code = get_cell(row_vals, col_map.get("code", -1)) if "code" in col_map else None
            serial = get_cell(row_vals, col_map.get("serial", -1)) if "serial" in col_map else None
            notes = get_cell(row_vals, col_map.get("notes", -1)) if "notes" in col_map else None
            image_ref = get_cell(row_vals, col_map.get("image_ref", -1)) if "image_ref" in col_map else None
            tutorial_url = get_cell(row_vals, col_map.get("tutorial_url", -1)) if "tutorial_url" in col_map else None

            pieces_qty = parse_pieces_qty(pieces_text)

            m = Material(
                lab_id=lab.id,
                name=name,
                location=location,
                status=status,
                pieces_text=pieces_text,
                pieces_qty=pieces_qty,
                brand=brand,
                model=model,
                code=code,
                serial=serial,
                notes=notes,
                image_ref=image_ref,
                tutorial_url=tutorial_url,
                source_file=os.path.basename(path),
                source_sheet=sheet_name,
                source_row=r,
            )

            db.session.add(m)
            inserted += 1

        # Commit por hoja
        db.session.commit()

    return inserted, skipped


def main():
    app = create_app()

    with app.app_context():
        if not os.path.isdir(INVENTORY_DIR):
            print(f"No existe carpeta: {INVENTORY_DIR}")
            print("Crea data/inventarios y pon ahí los Excel.")
            return

        files = [f for f in os.listdir(INVENTORY_DIR) if f.lower().endswith(".xlsx")]
        if not files:
            print("No hay archivos .xlsx en data/inventarios")
            return

        total_inserted = 0
        total_skipped = 0

        for f in files:
            path = os.path.join(INVENTORY_DIR, f)
            lab_name = derive_lab_name_from_filename(f)
            lab = ensure_lab(lab_name)

            print(f"\nImportando: {f} -> Lab: {lab.name}")
            inserted, skipped = import_workbook(path, lab)
            total_inserted += inserted
            total_skipped += skipped
            print(f"  Insertados: {inserted} | Saltados: {skipped}")

        print("\n=== RESUMEN ===")
        print("Total insertados:", total_inserted)
        print("Total saltados (filas sin nombre):", total_skipped)


if __name__ == "__main__":
    main()
