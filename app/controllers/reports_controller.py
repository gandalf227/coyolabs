import csv
import os
from datetime import datetime
from io import StringIO, BytesIO
from urllib.parse import urlencode
from flask import Blueprint, Response, current_app, render_template, request, url_for
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from app.models.lab import Lab
from app.models.material import Material
from app.models.debt import Debt
from app.models.inventory_request_ticket import InventoryRequestTicket
from app.models.critical_action_request import CriticalActionRequest
from app.models.logbook import LogbookEvent
from app.models.reservation import Reservation
from app.models.lost_found import LostFound
from app.models.software import Software
from app.utils.authz import min_role_required
from app.extensions import db

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


def csv_response(filename: str, headers: list[str], rows: list[list]):
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)

    data = buf.getvalue().encode("utf-8-sig")
    return Response(
        data,
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def excel_response(filename: str, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, _ in enumerate(headers, start=1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.font = header_font
        header_cell.alignment = header_alignment

    max_widths = [len(str(h)) if h is not None else 0 for h in headers]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = body_alignment
            text_value = "" if value is None else str(value)
            if len(text_value) > max_widths[col_idx - 1]:
                max_widths[col_idx - 1] = len(text_value)

    for col_idx, width in enumerate(max_widths, start=1):
        adjusted = max(12, min(width + 2, 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def parse_selected_columns(headers: list[str]) -> list[str]:
    selected = request.args.getlist("cols")
    if not selected:
        raw = (request.args.get("cols") or "").strip()
        if raw:
            selected = [part.strip() for part in raw.split(",") if part.strip()]
    if not selected:
        return headers
    allowed = set(headers)
    normalized = [col for col in selected if col in allowed]
    return normalized or headers


def project_rows(headers: list[str], rows: list[list], selected_columns: list[str]) -> tuple[list[str], list[list]]:
    index_by_name = {name: idx for idx, name in enumerate(headers)}
    projected_headers = [name for name in selected_columns if name in index_by_name]
    if not projected_headers:
        return headers, rows

    projected_rows = []
    for row in rows:
        projected_rows.append([row[index_by_name[name]] for name in projected_headers])
    return projected_headers, projected_rows


def build_download_url(endpoint: str) -> str:
    base = request.url_root.rstrip("/") + url_for(endpoint)
    pairs = []
    for key in request.args.keys():
        for value in request.args.getlist(key):
            pairs.append((key, value))
    query = urlencode(pairs, doseq=True)
    return f"{base}?{query}" if query else base


def build_inventory_rows(lab_id=None, status=None, search=None):
    q = Material.query
    if lab_id:
        q = q.filter(Material.lab_id == lab_id)
    if status:
        q = q.filter(Material.status == status)
    if search:
        q = q.filter(
            (Material.name.ilike(f"%{search}%"))
            | (Material.code.ilike(f"%{search}%"))
            | (Material.location.ilike(f"%{search}%"))
        )

    items = q.order_by(Material.lab_id, Material.location, Material.name).all()

    headers = [
        "id", "lab_id", "name", "location", "status",
        "pieces_text", "pieces_qty", "brand", "model", "code", "serial",
        "notes", "tutorial_url", "image_ref",
        "source_file", "source_sheet", "source_row",
        "created_at", "updated_at",
    ]

    rows = []
    for m in items:
        rows.append([
            m.id, m.lab_id, m.name, m.location, m.status,
            m.pieces_text, m.pieces_qty, m.brand, m.model, m.code, m.serial,
            m.notes, m.tutorial_url, m.image_ref,
            m.source_file, m.source_sheet, m.source_row,
            getattr(m, "created_at", None), getattr(m, "updated_at", None),
        ])
    return headers, rows


def build_debts_rows(status=None, user_id=None):
    q = Debt.query
    if status:
        q = q.filter(Debt.status == status)
    if user_id:
        q = q.filter(Debt.user_id == user_id)
    items = q.order_by(Debt.created_at.desc()).all()
    headers = ["id", "user_id", "material_id", "status", "reason", "amount", "created_at", "closed_at"]
    rows = []
    for d in items:
        rows.append([
            d.id, d.user_id, d.material_id, d.status, d.reason, d.amount, d.created_at, d.closed_at
        ])
    return headers, rows


def build_logbook_rows(action=None, module=None, user_id=None, material_id=None, description=None, date_from=None, date_to=None):
    q = LogbookEvent.query
    if action:
        q = q.filter(LogbookEvent.action.ilike(f"%{action}%"))
    if module:
        q = q.filter(LogbookEvent.module.ilike(f"%{module}%"))
    if user_id:
        q = q.filter(LogbookEvent.user_id == user_id)
    if material_id:
        q = q.filter(LogbookEvent.material_id == material_id)
    if description:
        q = q.filter(LogbookEvent.description.ilike(f"%{description}%"))
    if date_from:
        q = q.filter(func.date(LogbookEvent.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(LogbookEvent.created_at) <= date_to)

    items = q.order_by(LogbookEvent.created_at.desc()).all()
    headers = ["id", "user_id", "material_id", "module", "entity_label", "action", "description", "metadata_json", "created_at"]
    rows = []
    for e in items:
        rows.append([
            e.id, e.user_id, e.material_id, e.module, e.entity_label, e.action, e.description, e.metadata_json, e.created_at
        ])
    return headers, rows


def build_reservations_rows(status=None, room=None, user_id=None, date_from=None, date_to=None):
    q = Reservation.query
    if status:
        q = q.filter(Reservation.status == status)
    if room:
        q = q.filter(Reservation.room.ilike(f"%{room}%"))
    if user_id:
        q = q.filter(Reservation.user_id == user_id)
    if date_from:
        q = q.filter(Reservation.date >= date_from)
    if date_to:
        q = q.filter(Reservation.date <= date_to)

    items = q.order_by(Reservation.created_at.desc()).all()
    headers = [
        "id", "user_id", "room", "date", "start_time", "end_time", "status",
        "group_name", "teacher_name", "subject", "signed",
        "signature_ref", "admin_note", "purpose", "exit_time", "teacher_comments", "created_at",
    ]
    rows = []
    for r in items:
        rows.append([
            r.id, r.user_id, r.room, r.date, r.start_time, r.end_time, r.status,
            getattr(r, "group_name", None), getattr(r, "teacher_name", None), getattr(r, "subject", None),
            getattr(r, "signed", None),
            getattr(r, "signature_ref", None),
            r.admin_note, r.purpose, getattr(r, "exit_time", None), getattr(r, "teacher_comments", None),
            r.created_at,
        ])
    return headers, rows


def build_lostfound_rows():
    items = LostFound.query.order_by(LostFound.created_at.desc()).all()
    headers = [
        "id", "reported_by_user_id", "material_id", "title", "description",
        "location", "evidence_ref", "status", "admin_note", "created_at",
    ]
    rows = []
    for it in items:
        rows.append([
            it.id, it.reported_by_user_id, it.material_id, it.title, it.description,
            it.location, it.evidence_ref, it.status, it.admin_note, it.created_at,
        ])
    return headers, rows


def build_software_rows():
    items = Software.query.order_by(Software.name.asc()).all()
    headers = [
        "id", "lab_id", "name", "version", "license_type", "notes",
        "update_requested", "update_note", "created_at",
    ]
    rows = []
    for s in items:
        rows.append([
            s.id, s.lab_id, s.name, s.version, s.license_type, s.notes,
            s.update_requested, s.update_note, s.created_at,
        ])
    return headers, rows


def render_report_view(
    report_title,
    headers,
    rows,
    download_url,
    report_description=None,
    extra_meta=None,
    filter_fields=None,
    selected_columns=None,
    all_columns=None,
    download_excel_url=None,
    download_pdf_url=None,
):
    return render_template(
        "reports/report_view.html",
        report_title=report_title,
        columns=headers,
        rows=rows,
        all_columns=all_columns or headers,
        selected_columns=selected_columns or headers,
        download_url=download_url,
        download_excel_url=download_excel_url,
        download_pdf_url=download_pdf_url,
        report_description=report_description,
        extra_meta=extra_meta,
        filter_fields=filter_fields or [],
        active_page="reports",
    )


def _sanitize_pdf_cell(value) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.split())
    if len(text) > 120:
        return text[:117] + "..."
    return text or "-"


def pdf_response(
    *,
    filename: str,
    report_title: str,
    headers: list[str],
    rows: list[list],
    subtitle: str | None = None,
    download: bool = False,
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    story = []

    logo_path = os.path.join(current_app.root_path, "static", "img", "coyolabs.png")
    if os.path.exists(logo_path):
        story.append(Image(logo_path, width=0.8 * inch, height=0.8 * inch))
        story.append(Spacer(1, 6))

    story.append(Paragraph(report_title, styles["Title"]))
    if subtitle:
        story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    story.append(Spacer(1, 8))

    table_data = [headers] + [[_sanitize_pdf_cell(v) for v in row] for row in rows]
    available_width = doc.width
    col_count = max(1, len(headers))
    col_width = available_width / col_count
    table = Table(table_data, repeatRows=1, colWidths=[col_width] * col_count)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0E4C5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#5B4410")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D3C5A6")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)

    disposition = "attachment" if download else "inline"
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
    )


@reports_bp.route("/", methods=["GET"])
@min_role_required("ADMIN")
def reports_home():
    labs = Lab.query.order_by(Lab.name).all()

    reservations_by_status = (
        db.session.query(Reservation.status, func.count(Reservation.id))
        .group_by(Reservation.status)
        .order_by(func.count(Reservation.id).desc())
        .all()
    )

    room_usage = (
        db.session.query(Reservation.room, func.count(Reservation.id).label("total"))
        .group_by(Reservation.room)
        .order_by(func.count(Reservation.id).desc())
        .limit(10)
        .all()
    )

    inventory_daily_by_status = (
        db.session.query(InventoryRequestTicket.status, func.count(InventoryRequestTicket.id))
        .group_by(InventoryRequestTicket.status)
        .order_by(func.count(InventoryRequestTicket.id).desc())
        .all()
    )
    recent_inventory_tickets = (
        InventoryRequestTicket.query
        .order_by(InventoryRequestTicket.request_date.desc(), InventoryRequestTicket.created_at.desc())
        .limit(10)
        .all()
    )

    open_debts_count = Debt.query.filter(Debt.status == "OPEN").count()
    open_debts_by_user = (
        db.session.query(Debt.user_id, func.count(Debt.id).label("total"))
        .filter(Debt.status == "OPEN")
        .group_by(Debt.user_id)
        .order_by(func.count(Debt.id).desc())
        .limit(10)
        .all()
    )

    critical_actions_by_status = (
        db.session.query(CriticalActionRequest.status, func.count(CriticalActionRequest.id))
        .group_by(CriticalActionRequest.status)
        .order_by(func.count(CriticalActionRequest.id).desc())
        .all()
    )

    logbook_by_module = (
        db.session.query(func.coalesce(LogbookEvent.module, "SIN_MODULO"), func.count(LogbookEvent.id))
        .group_by(func.coalesce(LogbookEvent.module, "SIN_MODULO"))
        .order_by(func.count(LogbookEvent.id).desc())
        .all()
    )

    reservations_total = sum(total for _, total in reservations_by_status)
    inventory_daily_total = sum(total for _, total in inventory_daily_by_status)

    return render_template(
        "reports/home.html",
        labs=labs,
        reservations_by_status=reservations_by_status,
        room_usage=room_usage,
        inventory_daily_by_status=inventory_daily_by_status,
        recent_inventory_tickets=recent_inventory_tickets,
        open_debts_count=open_debts_count,
        open_debts_by_user=open_debts_by_user,
        critical_actions_by_status=critical_actions_by_status,
        logbook_by_module=logbook_by_module,
        reservations_total=reservations_total,
        inventory_daily_total=inventory_daily_total,
        active_page="reports",
    )


@reports_bp.route("/inventory.csv", methods=["GET"])
@min_role_required("ADMIN")
def report_inventory():
    lab_id = request.args.get("lab_id", type=int)
    status = (request.args.get("status") or "").strip()
    search = (request.args.get("search") or "").strip()
    headers, rows = build_inventory_rows(lab_id=lab_id, status=status or None, search=search or None)
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    fname = "inventory.csv" if not lab_id else f"inventory_lab_{lab_id}.csv"
    return csv_response(fname, headers, rows)


@reports_bp.route("/inventory.xlsx", methods=["GET"])
@min_role_required("ADMIN")
def report_inventory_excel():
    lab_id = request.args.get("lab_id", type=int)
    status = (request.args.get("status") or "").strip()
    search = (request.args.get("search") or "").strip()
    headers, rows = build_inventory_rows(lab_id=lab_id, status=status or None, search=search or None)
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    fname = "inventory.xlsx" if not lab_id else f"inventory_lab_{lab_id}.xlsx"
    return excel_response(fname, headers, rows)


@reports_bp.route("/view/inventory", methods=["GET"])
@min_role_required("ADMIN")
def report_inventory_view():
    lab_id = request.args.get("lab_id", type=int)
    status = (request.args.get("status") or "").strip()
    search = (request.args.get("search") or "").strip()
    headers, rows = build_inventory_rows(lab_id=lab_id, status=status or None, search=search or None)
    all_headers = headers[:]
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)

    report_title = "Inventario general"
    extra_meta = None
    if lab_id:
        lab = Lab.query.get(lab_id)
        report_title = f"Inventario - {lab.name if lab else f'Lab {lab_id}'}"
        extra_meta = f"Lab ID: {lab_id}"

    return render_report_view(
        report_title=report_title,
        headers=headers,
        rows=rows,
        download_url=build_download_url("reports.report_inventory"),
        report_description="Vista completa del inventario.",
        extra_meta=extra_meta,
        filter_fields=[
            {"name": "lab_id", "label": "Lab ID", "type": "number", "value": lab_id or "", "placeholder": "Ejemplo: 1"},
            {"name": "status", "label": "Estado", "type": "text", "value": status, "placeholder": "Ejemplo: DISPONIBLE"},
            {"name": "search", "label": "Buscar", "type": "text", "value": search, "placeholder": "Nombre, código o ubicación"},
        ],
        selected_columns=selected_columns,
        all_columns=all_headers,
        download_excel_url=build_download_url("reports.report_inventory_excel"),
        download_pdf_url=build_download_url("reports.report_inventory_pdf"),
    )


@reports_bp.route("/debts.csv", methods=["GET"])
@min_role_required("ADMIN")
def report_debts():
    status = (request.args.get("status") or "").strip()
    user_id = request.args.get("user_id", type=int)
    headers, rows = build_debts_rows(status=status or None, user_id=user_id)
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return csv_response("debts.csv", headers, rows)


@reports_bp.route("/debts.xlsx", methods=["GET"])
@min_role_required("ADMIN")
def report_debts_excel():
    status = (request.args.get("status") or "").strip()
    user_id = request.args.get("user_id", type=int)
    headers, rows = build_debts_rows(status=status or None, user_id=user_id)
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return excel_response("debts.xlsx", headers, rows)


@reports_bp.route("/view/debts", methods=["GET"])
@min_role_required("ADMIN")
def report_debts_view():
    status = (request.args.get("status") or "").strip()
    user_id = request.args.get("user_id", type=int)
    headers, rows = build_debts_rows(status=status or None, user_id=user_id)
    all_headers = headers[:]
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return render_report_view(
        report_title="Adeudos",
        headers=headers,
        rows=rows,
        download_url=build_download_url("reports.report_debts"),
        report_description="Vista completa de los adeudos registrados.",
        filter_fields=[
            {"name": "status", "label": "Estado", "type": "text", "value": status, "placeholder": "Ejemplo: OPEN"},
            {"name": "user_id", "label": "User ID", "type": "number", "value": user_id or "", "placeholder": "Ejemplo: 42"},
        ],
        selected_columns=selected_columns,
        all_columns=all_headers,
        download_excel_url=build_download_url("reports.report_debts_excel"),
        download_pdf_url=build_download_url("reports.report_debts_pdf"),
    )


@reports_bp.route("/logbook.csv", methods=["GET"])
@min_role_required("ADMIN")
def report_logbook():
    action = (request.args.get("action") or "").strip()
    module = (request.args.get("module") or "").strip()
    user_id = request.args.get("user_id", type=int)
    material_id = request.args.get("material_id", type=int)
    description = (request.args.get("description") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_logbook_rows(
        action=action or None,
        module=module or None,
        user_id=user_id,
        material_id=material_id,
        description=description or None,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return csv_response("logbook.csv", headers, rows)


@reports_bp.route("/logbook.xlsx", methods=["GET"])
@min_role_required("ADMIN")
def report_logbook_excel():
    action = (request.args.get("action") or "").strip()
    module = (request.args.get("module") or "").strip()
    user_id = request.args.get("user_id", type=int)
    material_id = request.args.get("material_id", type=int)
    description = (request.args.get("description") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_logbook_rows(
        action=action or None,
        module=module or None,
        user_id=user_id,
        material_id=material_id,
        description=description or None,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return excel_response("logbook.xlsx", headers, rows)


@reports_bp.route("/view/logbook", methods=["GET"])
@min_role_required("ADMIN")
def report_logbook_view():
    action = (request.args.get("action") or "").strip()
    module = (request.args.get("module") or "").strip()
    user_id = request.args.get("user_id", type=int)
    material_id = request.args.get("material_id", type=int)
    description = (request.args.get("description") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_logbook_rows(
        action=action or None,
        module=module or None,
        user_id=user_id,
        material_id=material_id,
        description=description or None,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    all_headers = headers[:]
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return render_report_view(
        report_title="Bitácora",
        headers=headers,
        rows=rows,
        download_url=build_download_url("reports.report_logbook"),
        report_description="Vista completa de la bitácora.",
        filter_fields=[
            {"name": "action", "label": "Acción contiene", "type": "text", "value": action, "placeholder": "Ejemplo: LOGIN"},
            {"name": "module", "label": "Módulo", "type": "text", "value": module, "placeholder": "Ejemplo: USERS"},
            {"name": "user_id", "label": "User ID", "type": "number", "value": user_id or "", "placeholder": "Ejemplo: 42"},
            {"name": "material_id", "label": "Material ID", "type": "number", "value": material_id or "", "placeholder": "Ejemplo: 128"},
            {"name": "description", "label": "Descripción contiene", "type": "text", "value": description, "placeholder": "Texto libre"},
            {"name": "date_from", "label": "Desde", "type": "date", "value": date_from},
            {"name": "date_to", "label": "Hasta", "type": "date", "value": date_to},
        ],
        selected_columns=selected_columns,
        all_columns=all_headers,
        download_excel_url=build_download_url("reports.report_logbook_excel"),
        download_pdf_url=build_download_url("reports.report_logbook_pdf"),
    )


@reports_bp.route("/reservations.csv", methods=["GET"])
@min_role_required("ADMIN")
def report_reservations():
    status = (request.args.get("status") or "").strip()
    room = (request.args.get("room") or "").strip()
    user_id = request.args.get("user_id", type=int)
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_reservations_rows(
        status=status or None,
        room=room or None,
        user_id=user_id,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return csv_response("reservations.csv", headers, rows)


@reports_bp.route("/reservations.xlsx", methods=["GET"])
@min_role_required("ADMIN")
def report_reservations_excel():
    status = (request.args.get("status") or "").strip()
    room = (request.args.get("room") or "").strip()
    user_id = request.args.get("user_id", type=int)
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_reservations_rows(
        status=status or None,
        room=room or None,
        user_id=user_id,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return excel_response("reservations.xlsx", headers, rows)


@reports_bp.route("/view/reservations", methods=["GET"])
@min_role_required("ADMIN")
def report_reservations_view():
    status = (request.args.get("status") or "").strip()
    room = (request.args.get("room") or "").strip()
    user_id = request.args.get("user_id", type=int)
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_reservations_rows(
        status=status or None,
        room=room or None,
        user_id=user_id,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    all_headers = headers[:]
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return render_report_view(
        report_title="Reservaciones",
        headers=headers,
        rows=rows,
        download_url=build_download_url("reports.report_reservations"),
        report_description="Vista completa de reservaciones.",
        filter_fields=[
            {"name": "status", "label": "Estado", "type": "text", "value": status, "placeholder": "Ejemplo: APPROVED"},
            {"name": "room", "label": "Salón/Lab", "type": "text", "value": room, "placeholder": "Ejemplo: LAB A"},
            {"name": "user_id", "label": "User ID", "type": "number", "value": user_id or "", "placeholder": "Ejemplo: 42"},
            {"name": "date_from", "label": "Desde", "type": "date", "value": date_from},
            {"name": "date_to", "label": "Hasta", "type": "date", "value": date_to},
        ],
        selected_columns=selected_columns,
        all_columns=all_headers,
        download_excel_url=build_download_url("reports.report_reservations_excel"),
        download_pdf_url=build_download_url("reports.report_reservations_pdf"),
    )


@reports_bp.route("/lostfound.csv", methods=["GET"])
@min_role_required("ADMIN")
def report_lostfound():
    headers, rows = build_lostfound_rows()
    return csv_response("lostfound.csv", headers, rows)


@reports_bp.route("/view/lostfound", methods=["GET"])
@min_role_required("ADMIN")
def report_lostfound_view():
    headers, rows = build_lostfound_rows()
    return render_report_view(
        report_title="Objetos perdidos",
        headers=headers,
        rows=rows,
        download_url=request.url_root.rstrip("/") + "/reports/lostfound.csv",
        report_description="Vista completa de objetos perdidos.",
    )


@reports_bp.route("/software.csv", methods=["GET"])
@min_role_required("ADMIN")
def report_software():
    headers, rows = build_software_rows()
    return csv_response("software.csv", headers, rows)


@reports_bp.route("/view/software", methods=["GET"])
@min_role_required("ADMIN")
def report_software_view():
    headers, rows = build_software_rows()
    return render_report_view(
        report_title="Software",
        headers=headers,
        rows=rows,
        download_url=request.url_root.rstrip("/") + "/reports/software.csv",
        report_description="Vista completa del software registrado.",
    )


@reports_bp.route("/logbook", methods=["GET"])
@min_role_required("ADMIN")
def logbook_admin_view():
    action = (request.args.get("action") or "").strip()
    module = (request.args.get("module") or "").strip()
    user_id = request.args.get("user_id", type=int)
    material_id = request.args.get("material_id", type=int)
    description = (request.args.get("description") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    headers, rows = build_logbook_rows(
        action=action or None,
        module=module or None,
        user_id=user_id,
        material_id=material_id,
        description=description or None,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    visible_columns, visible_rows = project_rows(headers, rows, selected_columns)
    export_url = build_download_url("reports.report_logbook")
    export_excel_url = build_download_url("reports.report_logbook_excel")
    export_pdf_url = build_download_url("reports.report_logbook_pdf")

    return render_template(
        "reports/logbook_admin.html",
        columns=visible_columns,
        rows=visible_rows[:500],
        all_columns=headers,
        selected_columns=selected_columns,
        export_url=export_url,
        export_excel_url=export_excel_url,
        export_pdf_url=export_pdf_url,
        action=action,
        module=module,
        user_id=user_id,
        material_id=material_id,
        description=description,
        date_from=date_from,
        date_to=date_to,
        active_page="reports",
    )


@reports_bp.route("/inventory.pdf", methods=["GET"])
@min_role_required("ADMIN")
def report_inventory_pdf():
    lab_id = request.args.get("lab_id", type=int)
    status = (request.args.get("status") or "").strip()
    search = (request.args.get("search") or "").strip()
    download = request.args.get("download", default=0, type=int) == 1
    headers, rows = build_inventory_rows(lab_id=lab_id, status=status or None, search=search or None)
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    filename = "inventory.pdf" if not lab_id else f"inventory_lab_{lab_id}.pdf"
    title = "Reporte de Inventario" if not lab_id else f"Reporte de Inventario - Lab {lab_id}"
    return pdf_response(
        filename=filename,
        report_title=title,
        headers=headers,
        rows=rows,
        subtitle="Inventario institucional",
        download=download,
    )


@reports_bp.route("/debts.pdf", methods=["GET"])
@min_role_required("ADMIN")
def report_debts_pdf():
    status = (request.args.get("status") or "").strip()
    user_id = request.args.get("user_id", type=int)
    download = request.args.get("download", default=0, type=int) == 1
    headers, rows = build_debts_rows(status=status or None, user_id=user_id)
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return pdf_response(
        filename="debts.pdf",
        report_title="Reporte de Adeudos",
        headers=headers,
        rows=rows,
        subtitle="Adeudos institucionales",
        download=download,
    )


@reports_bp.route("/logbook.pdf", methods=["GET"])
@min_role_required("ADMIN")
def report_logbook_pdf():
    action = (request.args.get("action") or "").strip()
    module = (request.args.get("module") or "").strip()
    user_id = request.args.get("user_id", type=int)
    material_id = request.args.get("material_id", type=int)
    description = (request.args.get("description") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    download = request.args.get("download", default=0, type=int) == 1

    headers, rows = build_logbook_rows(
        action=action or None,
        module=module or None,
        user_id=user_id,
        material_id=material_id,
        description=description or None,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return pdf_response(
        filename="logbook.pdf",
        report_title="Reporte de Bitácora",
        headers=headers,
        rows=rows,
        subtitle="Eventos administrativos y de operación",
        download=download,
    )


@reports_bp.route("/reservations.pdf", methods=["GET"])
@min_role_required("ADMIN")
def report_reservations_pdf():
    status = (request.args.get("status") or "").strip()
    room = (request.args.get("room") or "").strip()
    user_id = request.args.get("user_id", type=int)
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    download = request.args.get("download", default=0, type=int) == 1

    headers, rows = build_reservations_rows(
        status=status or None,
        room=room or None,
        user_id=user_id,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    selected_columns = parse_selected_columns(headers)
    headers, rows = project_rows(headers, rows, selected_columns)
    return pdf_response(
        filename="reservations.pdf",
        report_title="Reporte de Reservaciones",
        headers=headers,
        rows=rows,
        subtitle="Reservaciones institucionales",
        download=download,
    )
